import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ─── إعداد الصفحة ────────────────────────────────────────────────
st.set_page_config(
    page_title="Excel Filter Tool",
    page_icon="📊",
    layout="wide",
)

# ─── CSS مخصص ────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans+Arabic:wght@400;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'IBM Plex Sans Arabic', sans-serif;
    direction: rtl;
}

/* خلفية عامة */
.stApp { background-color: #1e1e2e; }

/* Sidebar */
[data-testid="stSidebar"] {
    background-color: #2a2a3e !important;
    border-left: 1px solid #3f3f5a;
}
[data-testid="stSidebar"] * { color: #e2e8f0 !important; }

/* العناوين */
h1 { color: #a5b4fc !important; font-size: 1.6rem !important; }
h3 { color: #a5b4fc !important; font-size: 1rem !important; }

/* الأزرار */
div.stButton > button {
    background-color: #7c83fd;
    color: #0f0f1a;
    font-weight: 700;
    border: none;
    border-radius: 6px;
    padding: 0.45rem 1.2rem;
    width: 100%;
    transition: background 0.2s;
}
div.stButton > button:hover { background-color: #a5b4fc; }

/* زر التصدير بالأوراق */
div[data-testid="stDownloadButton"] > button {
    background-color: #f59e0b !important;
    color: #0f0f1a !important;
    font-weight: 700 !important;
    border: none !important;
    border-radius: 6px !important;
    width: 100% !important;
}
div[data-testid="stDownloadButton"] > button:hover {
    background-color: #fbbf24 !important;
}

/* multiselect */
[data-baseweb="select"] {
    background-color: #2a2a3e !important;
    border-color: #3f3f5a !important;
}
[data-baseweb="tag"] { background-color: #4a4a8a !important; }

/* بطاقات المعلومات */
.info-card {
    background: #2a2a3e;
    border: 1px solid #3f3f5a;
    border-radius: 10px;
    padding: 14px 18px;
    margin-bottom: 10px;
    color: #e2e8f0;
}
.info-card .num {
    font-size: 2rem;
    font-weight: 700;
    color: #7c83fd;
    line-height: 1.1;
}
.info-card .lbl { font-size: 0.82rem; color: #94a3b8; }

/* شريط الفاصل */
hr { border-color: #3f3f5a !important; }

/* الجدول */
[data-testid="stDataFrame"] { border-radius: 8px; overflow: hidden; }

/* selectbox */
div[data-baseweb="select"] > div { background-color: #2a2a3e !important; }
</style>
""", unsafe_allow_html=True)


# ─── مساعدات ──────────────────────────────────────────────────────
def safe_sheet_name(name: str) -> str:
    for ch in r'\/*?[]:':
        name = name.replace(ch, "_")
    return name[:31]


def build_header_style():
    fill = PatternFill("solid", fgColor="FF4A4A8A")
    font = Font(bold=True, color="FFAAAAFF", name="Calibri")
    align = Alignment(horizontal="center", vertical="center")
    return fill, font, align


def export_single_sheet(df: pd.DataFrame) -> bytes:
    """تصدير ورقة واحدة تحتوي كل الصفوف المصفّاة."""
    wb = Workbook()
    ws = wb.active
    ws.title = "النتائج"
    fill, font, align = build_header_style()
    cols = list(df.columns)

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill, cell.font, cell.alignment = fill, font, align
        ws.column_dimensions[get_column_letter(ci)].width = max(12, len(str(col)) + 4)

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_multi_sheets(df: pd.DataFrame, col: str, keywords: list) -> bytes:
    """تصدير ورقة لكل قيمة فريدة."""
    wb = Workbook()
    wb.remove(wb.active)
    fill, font, align = build_header_style()
    df_cols = list(df.columns)

    total = len(keywords)

    for idx, kw in enumerate(keywords, 1):
        sheet_name = safe_sheet_name(str(kw))
        subset = df[df[col] == kw]
        ws = wb.create_sheet(title=sheet_name)

        for ci, c in enumerate(df_cols, 1):
            cell = ws.cell(row=1, column=ci, value=c)
            cell.fill, cell.font, cell.alignment = fill, font, align
            ws.column_dimensions[get_column_letter(ci)].width = max(12, len(str(c)) + 4)

        for ri, (_, row) in enumerate(subset.iterrows(), 2):
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── حالة الجلسة ──────────────────────────────────────────────────
if "df" not in st.session_state:
    st.session_state.df = None
if "filtered_df" not in st.session_state:
    st.session_state.filtered_df = None
if "multi_bytes" not in st.session_state:
    st.session_state.multi_bytes = None
if "saved_col" not in st.session_state:
    st.session_state.saved_col = None
if "saved_values" not in st.session_state:
    st.session_state.saved_values = []


# ══════════════════════════════════════════════════════════════════
#  الشريط الجانبي — الخطوات
# ══════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## 📊 Excel Filter Tool")
    st.markdown("---")

    # ── ① رفع الملف ──
    st.markdown("### ① رفع الملف")
    uploaded = st.file_uploader("اختر ملف Excel",
                                 type=["xlsx", "xls", "xlsm"],
                                 label_visibility="collapsed")

    if uploaded:
        try:
            df = pd.read_excel(uploaded, dtype=str).fillna("")
            st.session_state.df = df
            st.session_state.filtered_df = None
            st.success(f"✅ تم تحميل {len(df):,} صف")
        except Exception as e:
            st.error(f"فشل التحميل: {e}")

    st.markdown("---")

    # ── ② اختيار العمود ──
    st.markdown("### ② اختيار العمود")
    selected_col = None
    if st.session_state.df is not None:
        cols = list(st.session_state.df.columns)
        selected_col = st.selectbox("العمود", cols, label_visibility="collapsed")
    else:
        st.caption("ارفع ملف أولاً")

    st.markdown("---")

    # ── ③ اختيار القيم ──
    st.markdown("### ③ اختيار القيم")
    selected_values = []
    if st.session_state.df is not None and selected_col:
        unique_vals = sorted(
            st.session_state.df[selected_col].dropna().unique().tolist(),
            key=str
        )
        # بحث نصي سريع
        search = st.text_input("🔍 بحث في القيم", placeholder="اكتب للتصفية...",
                                label_visibility="collapsed")
        if search:
            unique_vals = [v for v in unique_vals if search.lower() in str(v).lower()]

        selected_values = st.multiselect(
            "القيم",
            options=unique_vals,
            label_visibility="collapsed",
            placeholder="اختر قيمة أو أكثر..."
        )

        c1, c2 = st.columns(2)
        if c1.button("تحديد الكل", use_container_width=True):
            st.session_state["_sel_all"] = unique_vals
            st.rerun()
        if c2.button("إلغاء الكل", use_container_width=True):
            st.session_state["_sel_all"] = []
            st.rerun()

        # تطبيق تحديد الكل / إلغاء الكل
        if "_sel_all" in st.session_state:
            selected_values = st.session_state.pop("_sel_all")

    st.markdown("---")

    # ── ④ تصفية ──
    st.markdown("### ④ تصفية وتصدير")
    if st.button("🔍  تصفية الصفوف", use_container_width=True):
        if st.session_state.df is None:
            st.warning("ارفع ملف أولاً")
        elif not selected_col:
            st.warning("اختر العمود")
        elif not selected_values:
            st.warning("اختر قيمة على الأقل")
        else:
            mask = st.session_state.df[selected_col].isin(selected_values)
            st.session_state.filtered_df = st.session_state.df[mask].copy()
            st.session_state.saved_col    = selected_col
            st.session_state.saved_values = selected_values
            st.session_state.multi_bytes  = None


# ══════════════════════════════════════════════════════════════════
#  المنطقة الرئيسية
# ══════════════════════════════════════════════════════════════════
st.markdown("# 📊 Excel Filter Tool")

df     = st.session_state.df
fdf    = st.session_state.filtered_df

# ── بطاقات الإحصاء ──
c1, c2, c3 = st.columns(3)
with c1:
    total_rows = len(df) if df is not None else 0
    st.markdown(f"""<div class="info-card">
        <div class="num">{total_rows:,}</div>
        <div class="lbl">إجمالي الصفوف</div></div>""", unsafe_allow_html=True)
with c2:
    filtered_rows = len(fdf) if fdf is not None else "—"
    st.markdown(f"""<div class="info-card">
        <div class="num">{filtered_rows if isinstance(filtered_rows,str) else f"{filtered_rows:,}"}</div>
        <div class="lbl">الصفوف المصفّاة</div></div>""", unsafe_allow_html=True)
with c3:
    n_vals = len(selected_values) if "selected_values" in dir() else 0
    st.markdown(f"""<div class="info-card">
        <div class="num">{n_vals}</div>
        <div class="lbl">قيم مختارة</div></div>""", unsafe_allow_html=True)

st.markdown("---")

def make_col_config(df: pd.DataFrame, sample_rows: int = 500) -> dict:
    """يحسب عرض كل عمود بناءً على أطول نص فيه."""
    config = {}
    sample = df.head(sample_rows)
    for c in df.columns:
        max_chars = max(
            len(str(c)),
            int(sample[c].astype(str).str.len().max()) if not sample.empty else 0
        )
        # كل حرف ≈ 8px، حد أدنى 80، حد أقصى 400
        px = max(80, min(400, max_chars * 8))
        config[c] = st.column_config.TextColumn(c, width=px)
    return config


# ── معاينة الجدول ──
if fdf is not None and not fdf.empty:
    st.markdown(f"### معاينة النتائج &nbsp; `{len(fdf):,} صف`")
    st.dataframe(fdf, column_config=make_col_config(fdf),
                 use_container_width=True, height=380)

    st.markdown("---")
    st.markdown("### ⬇️ تصدير")

    col_a, col_b = st.columns(2)

    with col_a:
        single_bytes = export_single_sheet(fdf)
        st.download_button(
            label="💾  تصدير (ورقة واحدة)",
            data=single_bytes,
            file_name="filtered_result.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    with col_b:
        _col    = st.session_state.saved_col
        _values = st.session_state.saved_values
        if _col and _values:
            multi_bytes = export_multi_sheets(fdf, _col, _values)
            st.download_button(
                label="📑  تصدير (ورقة لكل قيمة)",
                data=multi_bytes,
                file_name="sheets_result.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="dl_multi",
            )

elif df is not None:
    st.info("اختر القيم من الشريط الجانبي ثم اضغط **تصفية الصفوف**")
    preview = df.head(200)
    st.dataframe(preview, column_config=make_col_config(preview),
                 use_container_width=True, height=380)
else:
    st.markdown("""
    <div style="text-align:center; padding: 80px 0; color: #94a3b8;">
        <div style="font-size:4rem;">📂</div>
        <div style="font-size:1.2rem; margin-top:12px;">ارفع ملف Excel من الشريط الجانبي للبدء</div>
    </div>
    """, unsafe_allow_html=True)
 
